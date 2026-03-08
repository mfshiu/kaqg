# bank_loader.py
# 題庫載入器：從 _data/generation 讀取 xlsx，依科目代碼索引，啟動時全部載入記憶體。

import os
import random
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(os.getenv("LOGGER_NAME", __name__))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

# 題庫 xlsx：A1=總題數，第一題從第7列開始
BANK_FIRST_ROW = 7
BANK_CELL_TOTAL = "A1"
# 欄位：A=序號, B=試題內容, C-F=選項1-4, G=解答, H=章, I=節, J=頁碼, K=難度(可選), M=命名實體, N=文字化敘述
BANK_COL_STEM = 2
BANK_COL_OPT1, BANK_COL_OPT2, BANK_COL_OPT3, BANK_COL_OPT4 = 3, 4, 5, 6
BANK_COL_ANSWER = 7
BANK_COL_CHAPTER = 8
BANK_COL_SECTION = 9
BANK_COL_PAGE = 10
BANK_COL_DIFFICULTY = 11
BANK_COL_ENTITIES = 13   # M
BANK_COL_CLAUSES = 14    # N 文字化敘述


def _to_str(v) -> str:
    """將儲存格值轉為字串，避免 datetime 等型別呼叫 .strip() 失敗。"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _normalize_answer(a) -> str:
    if not a:
        return "1"
    a = _to_str(a).upper()
    mapping = {"A": "1", "1": "1", "B": "2", "2": "2", "C": "3", "3": "3", "D": "4", "4": "4"}
    return mapping.get(a, "1")


def load_bank_file(file_path: Path) -> list[dict]:
    """讀取單一題庫 xlsx，回傳 list of dict。使用 iter_rows 串流讀取。"""
    if load_workbook is None:
        raise ImportError("需要 openpyxl：python -m pip install openpyxl")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    try:
        total_val = ws[BANK_CELL_TOTAL].value
        total = int(total_val) if total_val is not None else 0
    except (TypeError, ValueError):
        wb.close()
        return []
    if total <= 0:
        wb.close()
        return []
    rows = []
    for row_cells in ws.iter_rows(min_row=BANK_FIRST_ROW, max_row=BANK_FIRST_ROW + total - 1, max_col=14, values_only=True):
        vals = list(row_cells) + [None] * (14 - len(row_cells))
        try:
            diff_val = vals[10] if len(vals) > 10 else None
            diff = int(diff_val) if diff_val is not None else None
        except (TypeError, ValueError):
            diff = None
        rows.append({
            "stem": _to_str(vals[1]) if len(vals) > 1 else "",
            "opt1": _to_str(vals[2]) if len(vals) > 2 else "",
            "opt2": _to_str(vals[3]) if len(vals) > 3 else "",
            "opt3": _to_str(vals[4]) if len(vals) > 4 else "",
            "opt4": _to_str(vals[5]) if len(vals) > 5 else "",
            "answer": _normalize_answer(vals[6] if len(vals) > 6 else None),
            "章": vals[7] if len(vals) > 7 else None,
            "節": vals[8] if len(vals) > 8 else None,
            "頁碼": vals[9] if len(vals) > 9 else None,
            "難度": diff,
            "entities": _to_str(vals[12]) if len(vals) > 12 else "",
            "clauses": _to_str(vals[13]) if len(vals) > 13 else "",
        })
    wb.close()
    return rows


class BankLoader:
    """題庫載入器：啟動時將 _data/generation 內所有 xlsx 載入，依科目代碼索引。"""

    def __init__(self, data_dir: Optional[str] = None):
        if data_dir is None:
            import app_helper
            data_dir = app_helper.get_generation_data_directory()
        self.data_dir = Path(data_dir)
        self._banks: dict[str, list[dict]] = {}
        self._load_all()

    def _load_all(self) -> None:
        """載入資料夾內所有 xlsx（檔名為科目代碼）。"""
        if not self.data_dir.exists() or not self.data_dir.is_dir():
            logger.warning("題庫目錄不存在：%s", self.data_dir)
            return
        for fp in sorted(self.data_dir.glob("*.xlsx")):
            if fp.name.startswith("~$"):
                continue
            subject = fp.stem
            try:
                rows = load_bank_file(fp)
                self._banks[subject] = rows
                logger.info("載入題庫 %s：%d 題", subject, len(rows))
            except Exception as e:
                logger.exception("載入題庫 %s 失敗：%s", subject, e)

    def get_subjects(self) -> list[str]:
        return list(self._banks.keys())

    def pick_template(self, subject: str, chapter: Optional[str] = None) -> Optional[dict]:
        """
        依科目、章隨機挑一樣板題。
        若章不存在則科目內隨機取一題。
        """
        rows = self._banks.get(subject)
        if not rows:
            return None
        if chapter is not None and str(chapter).strip():
            candidates = [r for r in rows if _chapter_match(r.get("章"), chapter)]
            if not candidates:
                candidates = rows
        else:
            candidates = rows
        return random.choice(candidates) if candidates else None


def _chapter_match(bank_chapter, requested: str) -> bool:
    """比對題庫的章與要求的章是否相符。"""
    if bank_chapter is None:
        return False
    return str(bank_chapter).strip() == str(requested).strip()
