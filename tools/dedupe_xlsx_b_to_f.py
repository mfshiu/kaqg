#!/usr/bin/env python3
"""Deduplicate rows in Excel files by columns B-F and reindex column A.

Usage:
  python tools/dedupe_xlsx_b_to_f.py --dir "D:\\Work\\ExamAI\\docs\\結案\\AI 3000題 v2"

Notes:
- Processes all .xlsx files recursively under --dir.
- Keeps the first occurrence of duplicated (B, C, D, E, F) values.
- Rewrites column A as 1..N after deduplication.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def normalize(value: Any) -> Any:
    """Normalize cell value for stable duplicate comparison."""
    if isinstance(value, str):
        return value.strip()
    return value


def iter_xlsx_files(root: Path) -> Iterable[Path]:
    for p in root.rglob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        yield p


def process_file(path: Path, dry_run: bool = False) -> tuple[int, int, int]:
    wb = load_workbook(path)
    total_removed = 0
    total_kept = 0

    for ws in wb.worksheets:
        max_row = ws.max_row
        if max_row < 2:
            continue

        seen: set[tuple[Any, Any, Any, Any, Any]] = set()
        keep_rows: list[tuple[Any, ...]] = []

        # Keep original header row
        header = tuple(ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1))

        for r in range(2, max_row + 1):
            key = tuple(normalize(ws.cell(row=r, column=c).value) for c in range(2, 7))
            if key in seen:
                total_removed += 1
                continue
            seen.add(key)
            row_values = tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))
            keep_rows.append(row_values)

        total_kept += len(keep_rows)

        if dry_run:
            continue

        # Rewrite sheet content: header + kept rows
        ws.delete_rows(2, ws.max_row)
        if keep_rows:
            for idx, row_values in enumerate(keep_rows, start=2):
                for c, v in enumerate(row_values, start=1):
                    ws.cell(row=idx, column=c, value=v)

        # Reindex column A starting from 1 for data rows
        for idx in range(2, ws.max_row + 1):
            ws.cell(row=idx, column=1, value=idx - 1)

        # Preserve header A as-is
        ws.cell(row=1, column=1, value=header[0] if header else ws.cell(row=1, column=1).value)

    if not dry_run:
        wb.save(path)

    return total_removed, total_kept, len(wb.worksheets)


def main() -> int:
    parser = argparse.ArgumentParser(description="Deduplicate xlsx rows by columns B-F and reindex column A")
    parser.add_argument("--dir", required=True, help="Target directory containing xlsx files")
    parser.add_argument("--dry-run", action="store_true", help="Only report what would change")
    args = parser.parse_args()

    root = Path(args.dir)
    if not root.exists() or not root.is_dir():
        raise SystemExit(f"Directory not found: {root}")

    files = list(iter_xlsx_files(root))
    if not files:
        print("No .xlsx files found.")
        return 0

    all_removed = 0
    all_kept = 0
    processed = 0

    for fp in files:
        removed, kept, sheets = process_file(fp, dry_run=args.dry_run)
        all_removed += removed
        all_kept += kept
        processed += 1
        action = "Would update" if args.dry_run else "Updated"
        print(f"{action}: {fp} | sheets={sheets} kept_rows={kept} removed_rows={removed}")

    print("---")
    print(f"Files processed: {processed}")
    print(f"Total kept rows: {all_kept}")
    print(f"Total removed rows: {all_removed}")
    print("Dry-run only." if args.dry_run else "Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
