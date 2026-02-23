#!/usr/bin/env python3
"""Rewrite duplicate question rows in xlsx files.

Features:
- Process all .xlsx files recursively under --dir.
- Target rows where stem in column B appears more than once in the same sheet.
- Shuffle options C/D/E/F and update answer in G accordingly (always write 1-4).
- Rewrite duplicated stems and options with an OpenAI model while preserving meaning.
- Emit progress logs to avoid appearing stuck.
- Write CSV change report for traceability.
"""

from __future__ import annotations

import argparse
import csv
import os
import random
import re
import time
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

from openai import OpenAI
from openpyxl import load_workbook

ANSWER_LETTERS = ["A", "B", "C", "D"]
ANSWER_COL = 7
OPTION_COLS = [3, 4, 5, 6]  # C, D, E, F
STEM_COL = 2


@dataclass
class ChangeRow:
    file: str
    sheet: str
    row: int
    stem_old: str
    stem_new: str
    options_old: str
    options_new: str
    answer_old: str
    answer_new: str
    notes: str


def normalize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.strip().split())
    return str(value).strip()


def similarity_ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def iter_xlsx_files(root: Path) -> Iterable[Path]:
    for p in root.rglob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        yield p


def parse_answer(answer_value: Any) -> str | None:
    text = normalize(answer_value).upper()
    if not text:
        return None
    m = re.search(r"[ABCD]", text)
    return m.group(0) if m else None


def parse_answer_index(answer_value: Any) -> int | None:
    """Return canonical answer index 0-3. Accepts A-D or 1-4."""
    text = normalize(answer_value).upper()
    if not text:
        return None

    num_match = re.search(r"[1-4]", text)
    if num_match:
        return int(num_match.group(0)) - 1

    letter = parse_answer(answer_value)
    if letter:
        return ANSWER_LETTERS.index(letter)
    return None


def format_answer_number(index: int) -> str:
    return str(index + 1)


def needs_shuffle(options: list[Any]) -> bool:
    normalized = [normalize(x) for x in options]
    return len(set(normalized)) > 1


def shuffled_indices(rng: random.Random) -> list[int]:
    idx = [0, 1, 2, 3]
    rng.shuffle(idx)
    return idx


def non_identity_shuffle(options: list[Any], rng: random.Random) -> list[int]:
    if not needs_shuffle(options):
        return [0, 1, 2, 3]
    for _ in range(20):
        idx = shuffled_indices(rng)
        if idx != [0, 1, 2, 3]:
            return idx
    return [1, 0, 2, 3]


def paraphrase_text(
    client: OpenAI,
    model: str,
    text: str,
    content_type: str,
    max_similarity: float,
    retry: int = 4,
) -> tuple[str, bool, float, int]:
    """Return rewritten text, diversity pass, ratio, and API call count."""
    kind_zh = "題幹" if content_type == "stem" else "選項"
    system_prompt = (
        "You are an expert editor for exam content in Traditional Chinese. "
        "Rewrite text with exactly the same meaning but clearly different wording and sentence structure. "
        "Keep domain terms accurate. Output only rewritten text."
    )
    user_prompt = f"""請改寫下列{kind_zh}，要求：
1. 語意必須完全一致，不可改變正確性或事實。
2. 用字與句型要明顯不同，盡量避免與原文長片段重複。
3. 保留專有名詞與關鍵術語的正確性。
4. 不要補充新資訊，不要解釋，不要加前後綴，只輸出改寫結果。

原文：
{text}"""

    last_error: Exception | None = None
    best_text = text
    best_ratio = 1.0
    api_calls = 0

    for attempt in range(1, retry + 1):
        try:
            api_calls += 1
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.95,
            )
            candidate = (response.choices[0].message.content or "").strip()
            if not candidate:
                time.sleep(0.8 * attempt)
                continue

            ratio = similarity_ratio(text, candidate)
            if normalize(candidate) != normalize(text) and ratio <= max_similarity:
                return candidate, True, ratio, api_calls

            if normalize(candidate) != normalize(text) and ratio < best_ratio:
                best_text = candidate
                best_ratio = ratio
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(1.2 * attempt)

    if last_error:
        raise RuntimeError(f"OpenAI paraphrase failed: {last_error}") from last_error
    if normalize(best_text) != normalize(text):
        return best_text, False, best_ratio, api_calls
    return text, False, 1.0, api_calls


def process_sheet(
    ws,
    client: OpenAI,
    model: str,
    rng: random.Random,
    dry_run: bool,
    progress_every: int,
    stem_max_similarity: float,
    option_max_similarity: float,
) -> tuple[list[ChangeRow], int]:
    changes: list[ChangeRow] = []
    api_calls = 0
    max_row = ws.max_row
    if max_row < 2:
        return changes, api_calls

    stem_rows: dict[str, list[int]] = {}
    for r in range(2, max_row + 1):
        stem = normalize(ws.cell(row=r, column=STEM_COL).value)
        if not stem:
            continue
        stem_rows.setdefault(stem, []).append(r)

    duplicate_groups = [rows for rows in stem_rows.values() if len(rows) >= 2]
    if not duplicate_groups:
        return changes, api_calls

    total_target_rows = sum(len(rows) for rows in duplicate_groups)
    stem_rewrite_target_rows = total_target_rows
    option_rewrite_target_rows = total_target_rows * 4
    print(
        f"  [Sheet {ws.title}] duplicate_groups={len(duplicate_groups)} "
        f"target_rows={total_target_rows} stem_rewrite_rows={stem_rewrite_target_rows} "
        f"option_rewrite_items={option_rewrite_target_rows}"
    )

    processed_rows = 0
    processed_stem_rewrite = 0
    processed_option_rewrite = 0

    for rows in duplicate_groups:
        for row_num in rows:
            stem_old = normalize(ws.cell(row=row_num, column=STEM_COL).value)
            options_old_values = [ws.cell(row=row_num, column=c).value for c in OPTION_COLS]
            answer_old_raw = ws.cell(row=row_num, column=ANSWER_COL).value
            answer_old_index = parse_answer_index(answer_old_raw)

            perm = non_identity_shuffle(options_old_values, rng)
            options_shuffled = [options_old_values[i] for i in perm]
            options_rewritten_values = list(options_shuffled)

            notes: list[str] = []
            answer_new_value: str | None = None
            if answer_old_index is not None:
                new_index = perm.index(answer_old_index)
                answer_new_value = format_answer_number(new_index)
            else:
                notes.append("answer_parse_failed")

            processed_stem_rewrite += 1
            print(
                f"    [API] sheet={ws.title} row={row_num} "
                f"stem ({processed_stem_rewrite}/{stem_rewrite_target_rows})..."
            )
            stem_new, stem_ok, stem_ratio, stem_calls = paraphrase_text(
                client=client,
                model=model,
                text=stem_old,
                content_type="stem",
                max_similarity=stem_max_similarity,
            )
            api_calls += stem_calls
            if not stem_ok:
                notes.append(f"stem_similarity_high:{stem_ratio:.3f}")

            for idx, opt in enumerate(options_shuffled, start=1):
                opt_text = normalize(opt)
                if not opt_text:
                    notes.append(f"option{idx}_empty")
                    continue
                processed_option_rewrite += 1
                if processed_option_rewrite % progress_every == 0 or processed_option_rewrite == option_rewrite_target_rows:
                    print(
                        f"    [API] sheet={ws.title} option "
                        f"{processed_option_rewrite}/{option_rewrite_target_rows}..."
                    )
                opt_new, opt_ok, opt_ratio, opt_calls = paraphrase_text(
                    client=client,
                    model=model,
                    text=opt_text,
                    content_type="option",
                    max_similarity=option_max_similarity,
                )
                api_calls += opt_calls
                options_rewritten_values[idx - 1] = opt_new
                if not opt_ok:
                    notes.append(f"option{idx}_similarity_high:{opt_ratio:.3f}")

            if not dry_run:
                ws.cell(row=row_num, column=STEM_COL, value=stem_new)
                for col, value in zip(OPTION_COLS, options_rewritten_values):
                    ws.cell(row=row_num, column=col, value=value)
                if answer_new_value is not None:
                    ws.cell(row=row_num, column=ANSWER_COL, value=answer_new_value)

            changes.append(
                ChangeRow(
                    file="",
                    sheet=ws.title,
                    row=row_num,
                    stem_old=stem_old,
                    stem_new=stem_new,
                    options_old=" | ".join(normalize(x) for x in options_old_values),
                    options_new=" | ".join(normalize(x) for x in options_rewritten_values),
                    answer_old=normalize(answer_old_raw),
                    answer_new=answer_new_value if answer_new_value is not None else normalize(answer_old_raw),
                    notes=",".join(notes),
                )
            )

            processed_rows += 1
            if processed_rows % progress_every == 0 or processed_rows == total_target_rows:
                print(f"    [Progress] sheet={ws.title} {processed_rows}/{total_target_rows} rows processed")

    return changes, api_calls


def process_file(
    path: Path,
    client: OpenAI,
    model: str,
    seed: int,
    dry_run: bool,
    progress_every: int,
    stem_max_similarity: float,
    option_max_similarity: float,
) -> tuple[list[ChangeRow], int]:
    wb = load_workbook(path)
    total_api_calls = 0
    file_changes: list[ChangeRow] = []

    for ws_index, ws in enumerate(wb.worksheets):
        rng = random.Random(seed + ws_index)
        changes, sheet_api_calls = process_sheet(
            ws,
            client=client,
            model=model,
            rng=rng,
            dry_run=dry_run,
            progress_every=max(1, progress_every),
            stem_max_similarity=stem_max_similarity,
            option_max_similarity=option_max_similarity,
        )
        for ch in changes:
            ch.file = str(path)
            file_changes.append(ch)
        total_api_calls += sheet_api_calls

    if file_changes and not dry_run:
        wb.save(path)

    return file_changes, total_api_calls


def write_report(report_path: Path, changes: list[ChangeRow]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "file",
                "sheet",
                "row",
                "stem_old",
                "stem_new",
                "options_old",
                "options_new",
                "answer_old",
                "answer_new",
                "notes",
            ]
        )
        for ch in changes:
            writer.writerow(
                [
                    ch.file,
                    ch.sheet,
                    ch.row,
                    ch.stem_old,
                    ch.stem_new,
                    ch.options_old,
                    ch.options_new,
                    ch.answer_old,
                    ch.answer_new,
                    ch.notes,
                ]
            )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Rewrite duplicated stems/options in xlsx files, shuffle options C-F, and update answer G"
    )
    parser.add_argument("--dir", required=True, help="Target directory containing xlsx files")
    parser.add_argument("--api-key", default=os.getenv("OPENAI_API_KEY"), help="OpenAI API key (or set OPENAI_API_KEY)")
    parser.add_argument("--model", default="gpt-4.1", help="OpenAI model for paraphrase")
    parser.add_argument("--seed", type=int, default=20260222, help="Random seed for option shuffling")
    parser.add_argument("--dry-run", action="store_true", help="Only report changes without writing xlsx")
    parser.add_argument("--progress-every", type=int, default=10, help="Print progress every N processed rows per sheet")
    parser.add_argument("--stem-max-similarity", type=float, default=0.80, help="Max allowed similarity for rewritten stems")
    parser.add_argument("--option-max-similarity", type=float, default=0.82, help="Max allowed similarity for rewritten options")
    parser.add_argument(
        "--report",
        default=f"_output/xlsx_duplicate_rewrite_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        help="CSV report output path",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    root = Path(args.dir)
    if not root.exists() or not root.is_dir():
        raise SystemExit(f"Directory not found: {root}")

    if not args.api_key:
        raise SystemExit("OpenAI API key is required. Use --api-key or set OPENAI_API_KEY.")

    files = list(iter_xlsx_files(root))
    if not files:
        print("No .xlsx files found.")
        return 0

    client = OpenAI(api_key=args.api_key)

    all_changes: list[ChangeRow] = []
    processed_files = 0
    api_calls = 0

    print(
        f"[Run] files={len(files)} model={args.model} "
        f"stem_max_similarity={args.stem_max_similarity} option_max_similarity={args.option_max_similarity}"
    )

    for fp in files:
        print(f"[File] Start: {fp}")
        file_changes, file_api_calls = process_file(
            fp,
            client=client,
            model=args.model,
            seed=args.seed,
            dry_run=args.dry_run,
            progress_every=args.progress_every,
            stem_max_similarity=args.stem_max_similarity,
            option_max_similarity=args.option_max_similarity,
        )
        all_changes.extend(file_changes)
        processed_files += 1
        api_calls += file_api_calls

        action = "Would update" if args.dry_run else "Updated"
        print(f"{action}: {fp} | changed_rows={len(file_changes)} | api_calls={file_api_calls}")
        for ch in file_changes[:20]:
            print(
                f"  - [{ch.sheet}] row={ch.row} answer {ch.answer_old} -> {ch.answer_new} "
                f"| stem_changed={'yes' if normalize(ch.stem_old) != normalize(ch.stem_new) else 'no'}"
            )
        if len(file_changes) > 20:
            print(f"  ... and {len(file_changes) - 20} more changed rows")

    report_path = Path(args.report)
    write_report(report_path, all_changes)

    print("---")
    print(f"Files processed: {processed_files}")
    print(f"Changed rows: {len(all_changes)}")
    print(f"API calls: {api_calls}")
    print(f"Report: {report_path}")
    print("Dry-run only." if args.dry_run else "Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
