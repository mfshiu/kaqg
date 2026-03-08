#!/usr/bin/env python3
"""從 AI_output 讀取 xlsx，每檔打亂後取前 300 題，重編序號，轉存至 extract 子目錄並依序改名。

用法：
  python tools/extract_sample_questions.py

路徑：
  輸入：D:\\Work\\kaqg\\_data\\H670\\AI_output
  輸出：D:\\Work\\kaqg\\_data\\H670\\AI_output\\extract
"""

from __future__ import annotations

import random
import time
from pathlib import Path

from openpyxl import load_workbook, Workbook


INPUT_DIR = Path(r"D:\Work\kaqg\_data\H670\AI_output")
OUTPUT_DIR = INPUT_DIR / "extract"

# 輸出檔名（依序對應第 1～15 個 xlsx）
OUTPUT_NAMES = [
    "1.廢棄物管理概論(甲乙丙級).xlsx",
    "2.專業技術人員職掌與工作倫理(甲乙丙級).xlsx",
    "3.廢棄物回收與再利用概論(乙丙級).xlsx",
    "4.廢棄物清理法(甲乙丙級).xlsx",
    "5.廢棄物清理許可及申報實務(甲乙丙級).xlsx",
    "6.廢棄物產源特性及減廢(乙級).xlsx",
    "7.廢棄物採樣檢測及特性分析(乙級).xlsx",
    "8.廢棄物貯存清除技術(乙丙級).xlsx",
    "9.廢棄物理化生物處理技術(乙級).xlsx",
    "10.廢棄物熱處理技術(乙級).xlsx",
    "11.廢棄物最終處置技術(乙級).xlsx",
    "12.廢棄物資源化與再利用技術(乙級).xlsx",
    "13.廢棄物貯存清除設備操作維護管理(乙丙級).xlsx",
    "14.廢棄物處理設施操作維護及營運管理(乙級).xlsx",
    "15.作業安全衛生及緊急應變(乙丙級).xlsx",
]

MAX_QUESTIONS = 300

# xlsx 欄位：序號=A(1), 試題內容=B(2), 選項1=C(3)..選項4=F(6), 解答=G(7), 章=H(8), 節=I(9), 頁碼=J(10), 難度=K(11)
COL_INDEX = 1
COL_STEM = 2
COL_OPT1 = 3
COL_OPT2 = 4
COL_OPT3 = 5
COL_OPT4 = 6
COL_ANSWER = 7
COL_CHAPTER = 8
COL_SECTION = 9
COL_PAGE = 10
COL_DIFFICULTY = 11

HEADERS = ["序號", "試題內容", "選項(1)", "選項(2)", "選項(3)", "選項(4)", "解答選項", "章", "節", "頁碼", "難度"]


def load_questions_from_xlsx(path: Path) -> list[dict]:
    """讀取 xlsx，回傳 list of {col: value}。使用 iter_rows 串流讀取，避免 cell() 逐格存取過慢。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row_cells in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        # row_cells: (A, B, C, D, E, F, G, H, I, J, K)
        vals = list(row_cells) + [None] * (11 - len(row_cells))  # 不足補 None
        rows.append({
            COL_INDEX: vals[0],
            COL_STEM: vals[1],
            COL_OPT1: vals[2],
            COL_OPT2: vals[3],
            COL_OPT3: vals[4],
            COL_OPT4: vals[5],
            COL_ANSWER: vals[6],
            COL_CHAPTER: vals[7],
            COL_SECTION: vals[8],
            COL_PAGE: vals[9],
            COL_DIFFICULTY: vals[10],
        })
    wb.close()
    return rows


def shuffle_and_take(rows: list[dict], max_n: int) -> list[dict]:
    """原序打亂後取前 max_n 筆。"""
    shuffled = list(rows)
    random.shuffle(shuffled)
    return shuffled[:max_n]


def save_questions(path: Path, rows: list[dict]) -> None:
    """寫入 xlsx，序號重編為 1..N。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(HEADERS)
    for idx, row in enumerate(rows, 1):
        ws.append([
            idx,
            row.get(COL_STEM, ""),
            row.get(COL_OPT1, ""),
            row.get(COL_OPT2, ""),
            row.get(COL_OPT3, ""),
            row.get(COL_OPT4, ""),
            row.get(COL_ANSWER, ""),
            row.get(COL_CHAPTER),
            row.get(COL_SECTION),
            row.get(COL_PAGE),
            row.get(COL_DIFFICULTY),
        ])
    wb.save(path)


def main() -> int:
    if not INPUT_DIR.exists() or not INPUT_DIR.is_dir():
        print(f"輸入目錄不存在：{INPUT_DIR}")
        return 1

    xlsx_files = sorted([p for p in INPUT_DIR.glob("*.xlsx") if not p.name.startswith("~$")])
    if not xlsx_files:
        print(f"目錄內無 xlsx 檔案：{INPUT_DIR}")
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"輸入目錄：{INPUT_DIR}")
    print(f"輸出目錄：{OUTPUT_DIR}")
    print(f"每檔打亂後取前 {MAX_QUESTIONS} 題，重編序號\n")

    total_elapsed = 0.0
    for i, src_path in enumerate(xlsx_files):
        out_name = OUTPUT_NAMES[i] if i < len(OUTPUT_NAMES) else f"{i + 1}.{src_path.name}"
        out_path = OUTPUT_DIR / out_name

        t0 = time.perf_counter()
        rows = load_questions_from_xlsx(src_path)
        t_load = time.perf_counter() - t0

        if not rows:
            print(f"  [{src_path.name}] 無資料，略過")
            continue

        t1 = time.perf_counter()
        sample = shuffle_and_take(rows, MAX_QUESTIONS)
        t_shuffle = time.perf_counter() - t1

        t2 = time.perf_counter()
        save_questions(out_path, sample)
        t_save = time.perf_counter() - t2

        total = t_load + t_shuffle + t_save
        total_elapsed += total
        print(f"  [{src_path.name}] -> [{out_name}] 原始 {len(rows)} 題 -> 取 {len(sample)} 題")
        print(f"      讀取: {t_load:.2f}s | 打亂: {t_shuffle:.2f}s | 寫入: {t_save:.2f}s | 合計: {total:.2f}s")

    print(f"\n完成。總耗時: {total_elapsed:.2f}s")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
