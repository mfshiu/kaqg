# 依序處理 _data/generation 內每個 xlsx：題幹+選項 → LLM 命名實體(第M欄) → KG 子句(第N欄)
# 需先啟動 kg_service、llm_service 與 MQTT broker。
# 用法（專案根目錄）：kg_name 由各 xlsx 檔名（不含副檔名）決定，不需輸入
#   python -m apps.xlsx_entities_clauses
# 若缺 openpyxl，請用「目前 Python」安裝（避免裝到系統）：
#   python -m pip install openpyxl

import json
import os
import sys
import time
import glob

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("缺少 openpyxl。請在目前環境執行：  python -m pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))
import app_helper
app_helper.initialize()

import logging
logger = logging.getLogger(os.getenv('LOGGER_NAME'))

from agentflow.core.agent import Agent
from agentflow.core.parcel import TextParcel

from services.kg_service import Topic as KgTopic
from services.llm_service import Topic as LlmTopic
from knowsys.knowledge_graph import KnowledgeGraph


# 試題欄位：第 7 列起為第一題，題幹=A, 選項1=B, 選項2=C, 選項3=D, 選項4=E
COL_STEM = 1    # A
COL_OPT1 = 2    # B
COL_OPT2 = 3    # C
COL_OPT3 = 4    # D
COL_OPT4 = 5    # E
COL_ENTITIES = 13   # M：命名實體（逗點分隔）
COL_CLAUSES = 14    # N：子句（分號分隔，主謂賓直接相連）
FIRST_ROW = 7
CELL_TOTAL = "A1"

NER_PROMPT_TEMPLATE = """請從以下試題文字中，抽出「命名實體」（專有名詞、重要概念、術語），不要選項代號或題幹中的 (A)(B)(C)(D)。
只回傳一個 JSON 物件，格式為：{"named_entities": ["實體1", "實體2", ...]}，不要其他說明。

試題文字：
---
%s
---"""


def extract_entities_from_llm_response(response_text):
    """從 LLM 回傳文字解析出 named_entities 列表。"""
    if response_text is None:
        return []
    if isinstance(response_text, dict):
        response_text = response_text.get("content") or response_text.get("response") or str(response_text)
    text = (response_text or "").strip()
    if not text:
        return []
    if "```" in text:
        parts = text.split("```")
        for p in parts:
            p = p.strip()
            if p.startswith("json"):
                p = p[4:].strip()
            if p.startswith("{"):
                try:
                    obj = json.loads(p)
                    return obj.get("named_entities", obj.get("entities", []))
                except json.JSONDecodeError:
                    continue
    try:
        obj = json.loads(text)
        return obj.get("named_entities", obj.get("entities", []))
    except json.JSONDecodeError:
        pass
    return []


def entities_to_clauses(kg, entity_list):
    """依命名實體列表查 KG：fact 節點 → 連出連入關聯 → 主謂賓子句（直接相連無符號），去重後以 list 回傳。"""
    clauses = []
    seen = set()
    for entity in (entity_list or []):
        entity = (entity or "").strip()
        if not entity:
            continue
        nodes = kg.query_nodes_by_name(entity, label="fact")
        if not nodes:
            nodes = kg.query_nodes_by_name(entity)
        for node in nodes:
            eid = node.get("element_id")
            if not eid:
                continue
            for subj, rel, obj in kg.query_all_relationships(eid):
                key = (subj, rel, obj)
                if key in seen:
                    continue
                seen.add(key)
                # 主+謂+賓直接相連，不用 —> 等符號
                clause = f"{subj} {rel} {obj}"
                clauses.append(clause)
    return clauses


class XlsxEntitiesClausesAgent(Agent):
    """對 xlsx 每題做 LLM NER 與 KG 子句查詢，寫入 M、N 欄。"""

    def __init__(self, agent_config, kg_name="kg01"):
        super().__init__(name="xlsx_entities_clauses", agent_config=agent_config)
        self.kg_name = kg_name
        self.bolt_url = None

    def ensure_kg_connection(self):
        if self.bolt_url:
            return True
        return_topic = self.agent_id
        self.subscribe(return_topic)
        pcl_ap = TextParcel({"kg_name": self.kg_name}, topic_return=return_topic)
        try:
            resp = self.publish_sync(KgTopic.ACCESS_POINT.value, pcl_ap, timeout=30)
        except TimeoutError:
            logger.error("KG ACCESS_POINT 逾時")
            return False
        content = resp.content if resp else {}
        self.bolt_url = content.get("bolt_url") if isinstance(content, dict) else None
        if not self.bolt_url:
            logger.error("未取得 bolt_url")
            return False
        return True

    def call_llm_ner(self, question_text):
        """題幹+選項文字 → LLM → 命名實體 list。"""
        return_topic = self.agent_id
        self.subscribe(return_topic)
        prompt = NER_PROMPT_TEMPLATE % (question_text or "").strip()
        pcl = TextParcel(
            {"messages": [{"role": "user", "content": prompt}]},
            topic_return=return_topic,
        )
        try:
            resp = self.publish_sync(LlmTopic.LLM_PROMPT.value, pcl, timeout=60)
        except TimeoutError:
            logger.warning("LLM 逾時")
            return []
        content = resp.content if resp else {}
        response_text = content.get("response") if isinstance(content, dict) else None
        return extract_entities_from_llm_response(response_text)

    def process_one_file(self, file_path):
        """處理單一 xlsx：檔名（不含副檔名）為 kg_name；讀 A1 總筆數，第 7 列起每題 LLM→M、KG→N，存檔。"""
        kg_name = os.path.splitext(os.path.basename(file_path))[0]
        if kg_name != self.kg_name:
            self.kg_name = kg_name
            self.bolt_url = None
        wb = load_workbook(file_path, read_only=False)
        ws = wb.active
        try:
            total_val = ws[CELL_TOTAL].value
            total = int(total_val) if total_val is not None else 0
        except (TypeError, ValueError):
            logger.warning("無法從 A1 讀取題目總筆數，跳過: %s", file_path)
            wb.close()
            return False
        if total <= 0:
            logger.info("題目總筆數為 0，跳過: %s", file_path)
            wb.close()
            return True

        if not self.ensure_kg_connection():
            wb.close()
            return False

        for i in range(total):
            row = FIRST_ROW + i
            stem = ws.cell(row=row, column=COL_STEM).value or ""
            o1 = ws.cell(row=row, column=COL_OPT1).value or ""
            o2 = ws.cell(row=row, column=COL_OPT2).value or ""
            o3 = ws.cell(row=row, column=COL_OPT3).value or ""
            o4 = ws.cell(row=row, column=COL_OPT4).value or ""
            question_text = f"{stem}\n(A) {o1}\n(B) {o2}\n(C) {o3}\n(D) {o4}"

            logger.info("處理 %s (kg=%s) 第 %d 題 (row %d)", os.path.basename(file_path), kg_name, i + 1, row)
            entities = self.call_llm_ner(question_text)
            entities_str = "，".join(entities) if entities else ""
            ws.cell(row=row, column=COL_ENTITIES, value=entities_str)

            clauses = []
            with KnowledgeGraph(uri=self.bolt_url) as kg:
                clauses = entities_to_clauses(kg, entities)
            clauses_str = "；".join(clauses) if clauses else ""
            ws.cell(row=row, column=COL_CLAUSES, value=clauses_str)

        wb.save(file_path)
        wb.close()
        logger.info("已存檔: %s", file_path)
        return True

    def run(self, xlsx_paths):
        """依序處理每個 xlsx。"""
        self.subscribe(self.agent_id)
        for path in xlsx_paths:
            self.process_one_file(path)
        return True


def main():
    data_dir = app_helper.get_generation_data_directory()
    pattern = os.path.join(data_dir, "*.xlsx")
    xlsx_files = sorted(glob.glob(pattern))
    if not xlsx_files:
        print(f"在 {data_dir} 未找到任何 xlsx 檔。")
        return

    print(f"找到 {len(xlsx_files)} 個 xlsx（kg_name 依檔名）")
    config = app_helper.get_agent_config()
    agent = XlsxEntitiesClausesAgent(config, kg_name="")
    agent.start_thread()
    time.sleep(1)
    agent.run(xlsx_files)
    time.sleep(1)
    agent.terminate()
    print("全部處理完成。")


if __name__ == "__main__":
    main()
