import random
import re
from turtle import st
import pdfplumber
import requests
import argparse
import time
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ------------------------------------------------------
# 從檔名自動解析起始頁碼：如 xxx-3.pdf → 3
# ------------------------------------------------------
def extract_start_page_from_filename(filename):
    try:
        base = os.path.splitext(filename)[0]
        parts = base.split("-")
        last = parts[-1]

        if last.isdigit():
            # 有尾端數字 → 視為起始頁
            clean_name = "-".join(parts[:-1]) if len(parts) > 1 else base
            return int(last), clean_name

        # 無尾端數字 → 起始頁 0
        return 0, base
    except Exception:
        # 任意解析錯誤 → 起始頁 0
        return 0, os.path.splitext(filename)[0]


# ------------------------------------------------------
# 呼叫 OSS GPT 出題（禁用「根據本文」等開頭）
# ------------------------------------------------------
def call_oss_gpt(text_chunk, difficulty):
    url = "http://140.115.53.67:11436/api/chat"

    prompt = f"""
你是一位專業教師，請依照以下規則出題：

【必須遵守的規則】
1. 題幹內容必須「僅」使用下列文字片段中的資訊，不可加入外部資料。
2. 題幹不得使用以下任何開頭：
   -「根據本文」
   -「根據上述文字」
   -「根據這段文字」
   -「依據本文」
   -「根據資料」
   -以及任何類似「引用前文」的句型。
   題幹必須直接敘述，不得引用文本來源。
3. 產生四個選項（1~4），且只能有一個正確答案。
4. 按照難度 {difficulty}（1=易、2=中、3=難）生成問題。

【文字片段】
{text_chunk}

請以以下 JSON 格式回覆（不要加入多餘文字）：
{{
  "stem": "...",
  "option1": "...",
  "option2": "...",
  "option3": "...",
  "option4": "...",
  "answer": "1|2|3|4"
}}
"""

    response = requests.post(
        url,
        json={
            "model": "gpt-oss:20b",
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "stream": False,
        }
    )

    result = response.json()
    return result["message"]["content"]

# ------------------------------------------------------
# PDF → 出題 → XLSX
# ------------------------------------------------------
def pdf_to_questions(pdf_path, start_page, xlsx_path):
    question_index = 1
    start_time = time.time()

    print("======================================")
    print(f"📘 處理 PDF: {pdf_path}")
    print(f"➡️ 起始頁：{start_page}")
    print(f"📊 輸出 XLSX: {xlsx_path}")
    print("======================================\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "序號", "試題內容", "選項(1)", "選項(2)",
        "選項(3)", "選項(4)", "解答選項",
        "章", "節", "頁碼", "難度"
    ]
    ws.append(headers)

    with pdfplumber.open(pdf_path) as pdf:

        for page_num in range(start_page - 1, len(pdf.pages)):
        # for page_num in range(start_page - 1, start_page+1):
            book_page = page_num + 2 - start_page
            print(f"\n--------------------------------------------------")
            print(f"📍 第 {book_page} 頁")

            page = pdf.pages[page_num]
            text = page.extract_text() or ""
            text = re.sub(r"\s+", "", text)
            text_len = len(text)

            print(f"   字數：{text_len}")

            if text_len < 100:
                print("   ⚠️ 跳過（不足 100 字）")
                continue

            chunk_size = 200
            chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
            print(f"   切成 {len(chunks)} 段")

            for i, chunk in enumerate(chunks):
                if len(chunk) <= chunk_size * 0.5:
                    break

                difficulty = random.choice([1, 2, 3])
                print(f"   ➡️ 段 {i+1}/{len(chunks)}，難度 {difficulty}")

                try:
                    resp_text = call_oss_gpt(chunk, difficulty)
                    q = eval(resp_text)
                    print(f"      ✔ 出題成功（題號 {question_index}）")
                except Exception as e:
                    print(f"      ✖ 出題失敗：{e}")
                    continue

                ws.append([
                    question_index,
                    q.get("stem", ""),
                    q.get("option1", ""),
                    q.get("option2", ""),
                    q.get("option3", ""),
                    q.get("option4", ""),
                    q.get("answer", ""),
                    "",          # 章
                    "",          # 節
                    book_page,
                    difficulty
                ])

                question_index += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(xlsx_path)

    elapsed = time.time() - start_time
    print("\n======================================")
    print("🎉 完成")
    print(f"📄 共 {question_index - 1} 題")
    print(f"📦 輸出：{xlsx_path}")
    print(f"⏱️ 耗時：{elapsed:.2f} 秒")
    print("======================================\n")


# ------------------------------------------------------
# 處理資料夾（多檔）
# ------------------------------------------------------
def process_folder(folder_path, default_start_page, output_folder):
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]

    print("\n======================================")
    print(f"📂 批次處理資料夾：{folder_path}")
    print(f"📄 找到 {len(pdf_files)} 個 PDF")
    print("======================================\n")

    os.makedirs(output_folder, exist_ok=True)

    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)

        # 從檔名解析 -n
        page_from_name, clean_name = extract_start_page_from_filename(pdf_file)
        start_page = page_from_name if page_from_name else default_start_page

        # 移除 -n 後的檔名
        xlsx_name = clean_name + ".xlsx"
        xlsx_path = os.path.join(output_folder, xlsx_name)

        print(f"📄 {pdf_file} → 📝 {xlsx_name}（起始頁：{start_page}）")

        pdf_to_questions(pdf_path, start_page, xlsx_path)


# ------------------------------------------------------
# CLI
# ------------------------------------------------------
def parse_args():
    parser = argparse.ArgumentParser(description="PDF 自動出題器")

    parser.add_argument("-p", "--pdf", help="處理單一 PDF 檔案")
    parser.add_argument("-d", "--dir", help="處理資料夾中所有 PDF")
    parser.add_argument("-s", "--start", type=int, default=0,
                        help="預設起始頁碼（若檔名含 -n 會覆蓋）")
    parser.add_argument("-o", "--out", help="輸出資料夾（預設同 PDF 位置）")

    return parser.parse_args()


# ------------------------------------------------------
# MAIN
# ------------------------------------------------------
if __name__ == "__main__":
    args = parse_args()

    # 單檔模式
    if args.pdf:
        start_page = args.start

        # 檢查檔名是否含 -n
        page_from_name, clean_name = extract_start_page_from_filename(os.path.basename(args.pdf))
        if page_from_name:
            start_page = page_from_name
            print(f"📌 由檔名解析到起始頁 {start_page}")

        output_folder = args.out if args.out else os.path.dirname(args.pdf)
        os.makedirs(output_folder, exist_ok=True)

        xlsx_path = os.path.join(output_folder, clean_name + ".xlsx")

        pdf_to_questions(args.pdf, start_page, xlsx_path)

    # 資料夾模式
    elif args.dir:
        output_folder = args.out if args.out else args.dir
        process_folder(args.dir, args.start, output_folder)

    else:
        print("❌ 請使用 -p 或 -d 其中一項")
