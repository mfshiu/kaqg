# 依 subject、數量、可選難度，從題庫 xlsx 隨機選樣板題+教材子句，用 LLM 生成新試題並輸出 xlsx。
# 需先啟動 llm_service 與 MQTT broker。
#
# 【試題生成失敗可能原因】
# 1. 題庫無題目：A1 總題數為 0 或無任何列可當樣板（缺難度/缺教材子句時會隨機補或由 LLM 依樣板出題）。
# 2. LLM 無回傳：逾時(90s)、llm_service 未啟動、MQTT 未連線或 API 錯誤。
# 3. LLM 回傳無法解析：回傳無法抽出有效 JSON 或缺少必要欄位；程式會儘量寬鬆解析（含 ```、結尾逗號、A/B/C/D 答案）。
# 4. 新題答案與樣板相同時不作變更，保持與 LLM 回傳相同。
# 失敗時日誌會印出「生成失敗原因」；若為解析失敗會另印「最後一次 LLM 回傳片段」供除錯（DEBUG 等級）。
#
# 【啟動較慢說明】程式啟動後會：讀取題庫 xlsx、連接 MQTT、再呼叫 LLM。首次呼叫 LLM 常需 10～30 秒
# （連線與 API 冷啟動），之後各題會較順。畫面上會印出進度，避免誤以為卡住。
#
# 用法（專案根目錄）：
#   python -m apps.gen_questions_from_bank <subject> <數量> [難度]
#   難度可省略，省略時難度與樣板題相同；指定時為 1=易 2=中 3=難
# 例：python -m apps.gen_questions_from_bank H67001 10     （難度同樣板）
#     python -m apps.gen_questions_from_bank H67001 10 2  （難度 2）
# 若缺 openpyxl：python -m pip install openpyxl

import json
import os
import random
import re
import sys
import time
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except ModuleNotFoundError:
    print("缺少 openpyxl。請在目前環境執行：  python -m pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))
import app_helper
app_helper.initialize()

import logging
logger = logging.getLogger(os.getenv('LOGGER_NAME'))

from agentflow.core.agent import Agent
from agentflow.core.parcel import TextParcel

from services.llm_service import Topic as LlmTopic


# 題庫 xlsx：A1=總題數，第一題從此列開始（若第7列為表頭、第8列才是第一題，請改為 8）
# 欄位依序：序號=A(1), 題幹=B(2), 選項1=C(3)..選項4=F(6), 解答=G(7), 章=H(8), 節=I(9), 頁碼=J(10), 難度=K(11), 教材子句=N(14)
BANK_FIRST_ROW = 7
BANK_CELL_TOTAL = "A1"
BANK_COL_INDEX = 1   # A 序號
BANK_COL_STEM = 2    # B 題幹
BANK_COL_OPT1 = 3    # C
BANK_COL_OPT2 = 4    # D
BANK_COL_OPT3 = 5    # E
BANK_COL_OPT4 = 6    # F
BANK_COL_ANSWER = 7  # G 解答選項
BANK_COL_CHAPTER = 8   # H 章
BANK_COL_SECTION = 9   # I 節
BANK_COL_PAGE = 10    # J 頁碼
BANK_COL_DIFFICULTY = 11  # K 難度
BANK_COL_CLAUSES = 14    # N 教材子句（分號分隔）

# 輸出 xlsx 表頭
OUTPUT_HEADERS = [
    "序號", "試題內容", "選項(1)", "選項(2)", "選項(3)", "選項(4)",
    "解答選項", "章", "節", "頁碼", "難度"
]

DIFFICULTY_NAMES = {1: "易", 2: "中", 3: "難"}

def _normalize_answer(a):
    """將解答正規為 "1"~"4"（題庫可能為 A/B/C/D 或 1/2/3/4）。"""
    if not a:
        return "1"
    a = str(a).strip().upper()
    if a in ("A", "1"):
        return "1"
    if a in ("B", "2"):
        return "2"
    if a in ("C", "3"):
        return "3"
    if a in ("D", "4"):
        return "4"
    return "1"

GEN_PROMPT_TEMPLATE = """你是一位專業教師。請根據「樣板題」的題型與「教材子句」的內容，產生一題新的選擇題。

【樣板題】
題幹：{stem}
選項1：{opt1}
選項2：{opt2}
選項3：{opt3}
選項4：{opt4}
正確答案：{answer}

【教材子句】
{clauses}

【要求】
1. 難度：{difficulty_name}（{difficulty}）。
2. 新題目的「正確答案」必須與樣板題不同（若樣板題答案為 1，新題答案須為 2、3 或 4）。
3. 若有教材子句，請僅依子句內容出題；若無教材子句，請依樣板題型與風格自行出題，內容需合理。
4. 回傳「一個」JSON 物件，不要其他說明。格式如下：
{{
  "stem": "新題幹",
  "option1": "選項1內容",
  "option2": "選項2內容",
  "option3": "選項3內容",
  "option4": "選項4內容",
  "answer": "1 或 2 或 3 或 4"
}}
"""
CLAUSES_PLACEHOLDER = "（無教材子句，請依樣板題型與風格自行出題，內容需合理且符合該難度。）"


def _extract_json_object(text):
    """從文字中取出第一個 {...} 並嘗試修掉常見 JSON 錯誤（如結尾多餘逗號）。"""
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                raw = text[start : i + 1]
                # 移除結尾多餘逗號，方便解析
                raw = re.sub(r",\s*}", "}", raw)
                raw = re.sub(r",\s*]", "]", raw)
                return raw
    return None


def parse_llm_question_json(response_text):
    """從 LLM 回傳解析出 stem, option1..4, answer；儘量寬鬆解析以產出試題。"""
    if response_text is None:
        return None
    if isinstance(response_text, dict):
        response_text = response_text.get("content") or response_text.get("response") or str(response_text)
    text = (response_text or "").strip()
    if not text:
        return None

    def make_result(obj):
        a = _normalize_answer(obj.get("answer"))
        return {
            "stem": (obj.get("stem") or obj.get("題幹") or "").strip(),
            "option1": (obj.get("option1") or obj.get("選項1") or "").strip(),
            "option2": (obj.get("option2") or obj.get("選項2") or "").strip(),
            "option3": (obj.get("option3") or obj.get("選項3") or "").strip(),
            "option4": (obj.get("option4") or obj.get("選項4") or "").strip(),
            "answer": a,
        }

    # 1) 嘗試 ```json ... ``` 區塊
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.lower().startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                try:
                    obj = json.loads(part)
                    return make_result(obj)
                except json.JSONDecodeError:
                    pass
            extracted = _extract_json_object(part)
            if extracted:
                try:
                    obj = json.loads(extracted)
                    return make_result(obj)
                except json.JSONDecodeError:
                    pass

    # 2) 整段直接解析
    try:
        obj = json.loads(text)
        return make_result(obj)
    except json.JSONDecodeError:
        pass

    # 3) 抽出第一個 {...} 再解析
    extracted = _extract_json_object(text)
    if extracted:
        try:
            obj = json.loads(extracted)
            return make_result(obj)
        except json.JSONDecodeError:
            pass
    return None


def load_question_bank(file_path):
    """讀取題庫 xlsx，回傳 list of dict（每題含 stem, opt1..4, answer, 章, 節, 頁碼, 難度, clauses）。"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    try:
        total_val = ws[BANK_CELL_TOTAL].value
        total = int(total_val) if total_val is not None else 0
    except (TypeError, ValueError):
        wb.close()
        return []
    if total <= 0:
        wb.close()
        return []
    rows = []
    for i in range(total):
        row_idx = BANK_FIRST_ROW + i
        try:
            diff_val = ws.cell(row=row_idx, column=BANK_COL_DIFFICULTY).value
            diff = int(diff_val) if diff_val is not None else None
        except (TypeError, ValueError):
            diff = None
        def _cell_str(row, col):
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ""

        raw_answer = _cell_str(row_idx, BANK_COL_ANSWER)
        # 章、節、頁碼讀自 H、I、J 欄，與題庫欄位定義一致（可能為數字，讀取時保留原值）
        chapter = ws.cell(row=row_idx, column=BANK_COL_CHAPTER).value
        section = ws.cell(row=row_idx, column=BANK_COL_SECTION).value
        page = ws.cell(row=row_idx, column=BANK_COL_PAGE).value
        rows.append({
            "stem": _cell_str(row_idx, BANK_COL_STEM),
            "opt1": _cell_str(row_idx, BANK_COL_OPT1),
            "opt2": _cell_str(row_idx, BANK_COL_OPT2),
            "opt3": _cell_str(row_idx, BANK_COL_OPT3),
            "opt4": _cell_str(row_idx, BANK_COL_OPT4),
            "answer": _normalize_answer(raw_answer),
            "章": chapter,
            "節": section,
            "頁碼": page,
            "難度": diff,
            "clauses": _cell_str(row_idx, BANK_COL_CLAUSES),
        })
    wb.close()
    return rows


class GenQuestionsAgent(Agent):
    """呼叫 LLM 依樣板題+教材子句生成新試題。"""

    def __init__(self, agent_config):
        super().__init__(name="gen_questions_from_bank", agent_config=agent_config)

    def call_llm_generate(self, prompt):
        """送 prompt 給 LLM，回傳 response 字串。"""
        return_topic = self.agent_id
        self.subscribe(return_topic)
        pcl = TextParcel(
            {"messages": [{"role": "user", "content": prompt}]},
            topic_return=return_topic,
        )
        try:
            resp = self.publish_sync(LlmTopic.LLM_PROMPT.value, pcl, timeout=90)
        except TimeoutError:
            logger.warning("LLM 逾時")
            return None
        content = resp.content if resp else {}
        return content.get("response") if isinstance(content, dict) else None


def generate_one_question(agent, bank_rows, difficulty=None, max_retries=3):
    """從 bank_rows 隨機選一樣板，呼叫 LLM 生成新題；新題答案須與樣板不同。
    缺難度時隨機 1/2/3；缺教材子句時仍可選用該樣板，由 LLM 依樣板風格出題。"""
    if difficulty is not None:
        # 指定難度：該難度或難度為空的題目都可當候選（難度空者以指定難度計）
        candidates = [r for r in bank_rows if r.get("難度") == difficulty or r.get("難度") is None]
    else:
        candidates = list(bank_rows)
    if not candidates:
        logger.warning("題庫中無任何題目可作樣板")
        return None
    template = random.choice(candidates)
    # 難度：樣板有則用樣板，否則若使用者有指定用指定值，否則隨機 1/2/3
    if template.get("難度") in (1, 2, 3):
        use_difficulty = template.get("難度")
    elif difficulty is not None:
        use_difficulty = difficulty
    else:
        use_difficulty = random.choice([1, 2, 3])
    template_answer = (template.get("answer") or "1").strip()
    if template_answer not in ("1", "2", "3", "4"):
        template_answer = "1"
    difficulty_name = DIFFICULTY_NAMES.get(use_difficulty, str(use_difficulty))
    clauses_text = (template.get("clauses") or "").strip()
    if not clauses_text:
        clauses_text = CLAUSES_PLACEHOLDER
    prompt = GEN_PROMPT_TEMPLATE.format(
        stem=template.get("stem", ""),
        opt1=template.get("opt1", ""),
        opt2=template.get("opt2", ""),
        opt3=template.get("opt3", ""),
        opt4=template.get("opt4", ""),
        answer=template_answer,
        clauses=clauses_text,
        difficulty=use_difficulty,
        difficulty_name=difficulty_name,
    )
    last_fail_reason = None
    last_response_snippet = None
    for attempt in range(max_retries):
        response = agent.call_llm_generate(prompt)
        if response is None:
            last_fail_reason = "LLM 無回傳（逾時或連線失敗）"
            last_response_snippet = None
            continue
        parsed = parse_llm_question_json(response)
        if not parsed:
            last_fail_reason = "LLM 回傳無法解析為 JSON 或缺少 stem/option1~4/answer"
            text = response if isinstance(response, str) else str(response)
            last_response_snippet = (text[:200] + "…") if len(text) > 200 else text
            continue
        new_answer = parsed.get("answer", "1")
        # 若新題答案與樣板相同，不作任何變更，保持相同
        # 章、節、頁碼必須與樣板題相同，直接取自樣板（不從 LLM 回傳）
        chapter = template.get("章")
        section = template.get("節")
        page = template.get("頁碼")
        return {
            "stem": parsed.get("stem", ""),
            "option1": parsed.get("option1", ""),
            "option2": parsed.get("option2", ""),
            "option3": parsed.get("option3", ""),
            "option4": parsed.get("option4", ""),
            "answer": new_answer,
            "章": chapter,
            "節": section,
            "頁碼": page,
            "難度": use_difficulty,
        }
    logger.warning("生成失敗原因：%s", last_fail_reason)
    if last_response_snippet:
        logger.debug("最後一次 LLM 回傳片段：%s", last_response_snippet)
    return None


def main():
    if len(sys.argv) < 3:
        print("用法：python -m apps.gen_questions_from_bank <subject> <數量> [難度]")
        print("  難度可省略，省略時難度與樣板題相同；若指定則為 1=易 2=中 3=難")
        sys.exit(1)
    subject = sys.argv[1].strip()
    try:
        count = int(sys.argv[2])
    except ValueError:
        print("數量請填數字。")
        sys.exit(1)
    difficulty = None
    if len(sys.argv) >= 4:
        try:
            difficulty = int(sys.argv[3])
            if difficulty not in (1, 2, 3):
                print("難度須為 1、2 或 3。")
                sys.exit(1)
        except ValueError:
            print("難度請填數字 1、2 或 3。")
            sys.exit(1)
    if count <= 0:
        print("數量須 > 0。")
        sys.exit(1)

    data_dir = app_helper.get_generation_data_directory()
    bank_path = os.path.join(data_dir, f"{subject}.xlsx")
    if not os.path.isfile(bank_path):
        print(f"題庫不存在：{bank_path}")
        sys.exit(1)

    print("正在讀取題庫…", flush=True)
    bank_rows = load_question_bank(bank_path)
    if not bank_rows:
        print("題庫無題目或 A1 總題數無效。")
        sys.exit(1)
    if difficulty is not None:
        print(f"題庫共 {len(bank_rows)} 題，欲生成 {count} 題，難度 {difficulty}（{DIFFICULTY_NAMES[difficulty]}）。")
    else:
        print(f"題庫共 {len(bank_rows)} 題，欲生成 {count} 題，難度與樣板題相同。")

    print("正在連接 MQTT 與啟動 agent…", flush=True)
    config = app_helper.get_agent_config()
    agent = GenQuestionsAgent(config)
    agent.start_thread()
    time.sleep(1)
    agent.subscribe(agent.agent_id)
    print("開始生成試題（首次呼叫 LLM 可能需 10～30 秒，屬正常）…", flush=True)

    results = []
    for i in range(count):
        print(f"  第 {i + 1}/{count} 題…", flush=True)
        logger.info("生成第 %d/%d 題", i + 1, count)
        row = generate_one_question(agent, bank_rows, difficulty)
        if row is None:
            logger.warning("第 %d 題生成失敗，略過（見上方「生成失敗原因」）", i + 1)
            continue
        results.append(row)

    agent.terminate()
    time.sleep(1)

    output_dir = os.path.join(os.path.dirname(app_helper.get_config_path()), "_output")
    os.makedirs(output_dir, exist_ok=True)
    diff_suffix = str(difficulty) if difficulty is not None else "same"
    out_name = f"{subject}_gen_{count}_{diff_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(OUTPUT_HEADERS)
    for idx, r in enumerate(results, 1):
        # 章、節、頁碼、難度皆與樣板題相同，來自 generate_one_question 的 template
        ws.append([
            idx,
            r.get("stem", ""),
            r.get("option1", ""),
            r.get("option2", ""),
            r.get("option3", ""),
            r.get("option4", ""),
            r.get("answer", ""),
            r.get("章"),   # 與樣板題相同
            r.get("節"),   # 與樣板題相同
            r.get("頁碼"), # 與樣板題相同
            r.get("難度"),
        ])
    wb.save(out_path)
    print(f"已寫入 {len(results)} 題至 {out_path}")


if __name__ == "__main__":
    main()
