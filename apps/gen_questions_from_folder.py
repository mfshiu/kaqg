# 依題庫 xlsx 由上至下逐題處理，用 LLM 生成新試題並輸出。
# 處理指定資料夾內所有 xlsx，結果輸出至該資料夾的 AI_output 子目錄。
# 需先啟動 llm_service 與 MQTT broker。
#
# 用法（專案根目錄）：
#   python -m apps.gen_questions_from_folder
#
# 預設路徑：
#   輸入：D:\Work\ExamAI\docs\廢棄物文件\廢棄物題庫\H670
#   輸出：D:\Work\ExamAI\docs\廢棄物文件\廢棄物題庫\H670\AI_output

import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except ModuleNotFoundError:
    print("缺少 openpyxl。請在目前環境執行：  python -m pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))
import app_helper
app_helper.initialize()

import logging
logger = logging.getLogger(os.getenv('LOGGER_NAME'))

from agentflow.core.agent import Agent
from agentflow.core.parcel import TextParcel

from services.llm_service import Topic as LlmTopic


# 路徑設定
INPUT_DIR = r"D:\Work\kaqg\_data\H670"
OUTPUT_DIR = os.path.join(INPUT_DIR, "AI_output")

# 題庫 xlsx：A1=總題數，第一題從此列開始
BANK_FIRST_ROW = 7
BANK_CELL_TOTAL = "A1"
BANK_COL_INDEX = 1   # A 序號
BANK_COL_STEM = 2    # B 題幹
BANK_COL_OPT1 = 3    # C
BANK_COL_OPT2 = 4    # D
BANK_COL_OPT3 = 5    # E
BANK_COL_OPT4 = 6    # F
BANK_COL_ANSWER = 7  # G 解答選項
BANK_COL_CHAPTER = 8   # H 章
BANK_COL_SECTION = 9   # I 節
BANK_COL_PAGE = 10    # J 頁碼
BANK_COL_DIFFICULTY = 11  # K 難度
BANK_COL_CLAUSES = 14    # N 教材子句（分號分隔）

# 輸出 xlsx 表頭
OUTPUT_HEADERS = [
    "序號", "試題內容", "選項(1)", "選項(2)", "選項(3)", "選項(4)",
    "解答選項", "章", "節", "頁碼", "難度"
]

DIFFICULTY_NAMES = {1: "易", 2: "中", 3: "難"}


def _normalize_answer(a):
    """將解答正規為 "1"~"4"（題庫可能為 A/B/C/D 或 1/2/3/4）。"""
    if not a:
        return "1"
    a = str(a).strip().upper()
    if a in ("A", "1"):
        return "1"
    if a in ("B", "2"):
        return "2"
    if a in ("C", "3"):
        return "3"
    if a in ("D", "4"):
        return "4"
    return "1"


GEN_PROMPT_TEMPLATE = """你是一位專業教師。請根據「樣板題」的題型與「教材子句」的內容，產生一題新的選擇題。

【樣板題】
題幹：{stem}
選項1：{opt1}
選項2：{opt2}
選項3：{opt3}
選項4：{opt4}
正確答案：{answer}

【教材子句】
{clauses}

【要求】
1. 難度：{difficulty_name}（{difficulty}）。
2. 新題目的「正確答案」必須與樣板題不同（若樣板題答案為 1，新題答案須為 2、3 或 4）。
3. 若有教材子句，請僅依子句內容出題；若無教材子句，請依樣板題型與風格自行出題，內容需合理。
4. 回傳「一個」JSON 物件，不要其他說明。格式如下：
{{
  "stem": "新題幹",
  "option1": "選項1內容",
  "option2": "選項2內容",
  "option3": "選項3內容",
  "option4": "選項4內容",
  "answer": "1 或 2 或 3 或 4"
}}
"""
CLAUSES_PLACEHOLDER = "（無教材子句，請依樣板題型與風格自行出題，內容需合理且符合該難度。）"


def _extract_json_object(text):
    """從文字中取出第一個 {...} 並嘗試修掉常見 JSON 錯誤（如結尾多餘逗號）。"""
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                raw = text[start : i + 1]
                raw = re.sub(r",\s*}", "}", raw)
                raw = re.sub(r",\s*]", "]", raw)
                return raw
    return None


def parse_llm_question_json(response_text):
    """從 LLM 回傳解析出 stem, option1..4, answer；儘量寬鬆解析以產出試題。"""
    if response_text is None:
        return None
    if isinstance(response_text, dict):
        response_text = response_text.get("content") or response_text.get("response") or str(response_text)
    text = (response_text or "").strip()
    if not text:
        return None

    def make_result(obj):
        a = _normalize_answer(obj.get("answer"))
        return {
            "stem": (obj.get("stem") or obj.get("題幹") or "").strip(),
            "option1": (obj.get("option1") or obj.get("選項1") or "").strip(),
            "option2": (obj.get("option2") or obj.get("選項2") or "").strip(),
            "option3": (obj.get("option3") or obj.get("選項3") or "").strip(),
            "option4": (obj.get("option4") or obj.get("選項4") or "").strip(),
            "answer": a,
        }

    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.lower().startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                try:
                    obj = json.loads(part)
                    return make_result(obj)
                except json.JSONDecodeError:
                    pass
            extracted = _extract_json_object(part)
            if extracted:
                try:
                    obj = json.loads(extracted)
                    return make_result(obj)
                except json.JSONDecodeError:
                    pass

    try:
        obj = json.loads(text)
        return make_result(obj)
    except json.JSONDecodeError:
        pass

    extracted = _extract_json_object(text)
    if extracted:
        try:
            obj = json.loads(extracted)
            return make_result(obj)
        except json.JSONDecodeError:
            pass
    return None


def load_question_bank(file_path):
    """讀取題庫 xlsx，回傳 list of dict（每題含 stem, opt1..4, answer, 章, 節, 頁碼, 難度, clauses）。"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    try:
        total_val = ws[BANK_CELL_TOTAL].value
        total = int(total_val) if total_val is not None else 0
    except (TypeError, ValueError):
        wb.close()
        return []
    if total <= 0:
        wb.close()
        return []
    rows = []
    for i in range(total):
        row_idx = BANK_FIRST_ROW + i
        try:
            diff_val = ws.cell(row=row_idx, column=BANK_COL_DIFFICULTY).value
            diff = int(diff_val) if diff_val is not None else None
        except (TypeError, ValueError):
            diff = None
        def _cell_str(row, col):
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ""

        raw_answer = _cell_str(row_idx, BANK_COL_ANSWER)
        chapter = ws.cell(row=row_idx, column=BANK_COL_CHAPTER).value
        section = ws.cell(row=row_idx, column=BANK_COL_SECTION).value
        page = ws.cell(row=row_idx, column=BANK_COL_PAGE).value
        rows.append({
            "stem": _cell_str(row_idx, BANK_COL_STEM),
            "opt1": _cell_str(row_idx, BANK_COL_OPT1),
            "opt2": _cell_str(row_idx, BANK_COL_OPT2),
            "opt3": _cell_str(row_idx, BANK_COL_OPT3),
            "opt4": _cell_str(row_idx, BANK_COL_OPT4),
            "answer": _normalize_answer(raw_answer),
            "章": chapter,
            "節": section,
            "頁碼": page,
            "難度": diff,
            "clauses": _cell_str(row_idx, BANK_COL_CLAUSES),
        })
    wb.close()
    return rows


class GenQuestionsAgent(Agent):
    """呼叫 LLM 依樣板題+教材子句生成新試題。"""

    def __init__(self, agent_config):
        super().__init__(name="gen_questions_from_folder", agent_config=agent_config)

    def call_llm_generate(self, prompt):
        """送 prompt 給 LLM，回傳 response 字串。"""
        return_topic = self.agent_id
        self.subscribe(return_topic)
        pcl = TextParcel(
            {"messages": [{"role": "user", "content": prompt}]},
            topic_return=return_topic,
        )
        try:
            resp = self.publish_sync(LlmTopic.LLM_PROMPT.value, pcl, timeout=90)
        except TimeoutError:
            logger.warning("LLM 逾時")
            return None
        content = resp.content if resp else {}
        return content.get("response") if isinstance(content, dict) else None


def generate_one_question(agent, template, max_retries=3):
    """以 template 為樣板，呼叫 LLM 生成新題；新題答案須與樣板不同。"""
    use_difficulty = template.get("難度")
    if use_difficulty not in (1, 2, 3):
        use_difficulty = 1
    template_answer = (template.get("answer") or "1").strip()
    if template_answer not in ("1", "2", "3", "4"):
        template_answer = "1"
    difficulty_name = DIFFICULTY_NAMES.get(use_difficulty, str(use_difficulty))
    clauses_text = (template.get("clauses") or "").strip()
    if not clauses_text:
        clauses_text = CLAUSES_PLACEHOLDER
    prompt = GEN_PROMPT_TEMPLATE.format(
        stem=template.get("stem", ""),
        opt1=template.get("opt1", ""),
        opt2=template.get("opt2", ""),
        opt3=template.get("opt3", ""),
        opt4=template.get("opt4", ""),
        answer=template_answer,
        clauses=clauses_text,
        difficulty=use_difficulty,
        difficulty_name=difficulty_name,
    )
    last_fail_reason = None
    last_response_snippet = None
    for attempt in range(max_retries):
        response = agent.call_llm_generate(prompt)
        if response is None:
            last_fail_reason = "LLM 無回傳（逾時或連線失敗）"
            last_response_snippet = None
            continue
        parsed = parse_llm_question_json(response)
        if not parsed:
            last_fail_reason = "LLM 回傳無法解析為 JSON 或缺少 stem/option1~4/answer"
            text = response if isinstance(response, str) else str(response)
            last_response_snippet = (text[:200] + "…") if len(text) > 200 else text
            continue
        new_answer = parsed.get("answer", "1")
        chapter = template.get("章")
        section = template.get("節")
        page = template.get("頁碼")
        return {
            "stem": parsed.get("stem", ""),
            "option1": parsed.get("option1", ""),
            "option2": parsed.get("option2", ""),
            "option3": parsed.get("option3", ""),
            "option4": parsed.get("option4", ""),
            "answer": new_answer,
            "章": chapter,
            "節": section,
            "頁碼": page,
            "難度": use_difficulty,
        }
    logger.warning("生成失敗原因：%s", last_fail_reason)
    if last_response_snippet:
        logger.debug("最後一次 LLM 回傳片段：%s", last_response_snippet)
    return None


def process_one_file(agent, file_path, output_dir):
    """處理單一 xlsx，由上至下逐題生成，輸出至 output_dir。"""
    bank_rows = load_question_bank(file_path)
    if not bank_rows:
        logger.warning("題庫無題目或 A1 總題數無效：%s", file_path)
        return 0
    base_name = Path(file_path).stem
    total = len(bank_rows)
    results = []
    for i, template in enumerate(bank_rows):
        clauses = (template.get("clauses") or "").strip()
        clauses_preview = (clauses[:100] + "…") if len(clauses) > 100 else clauses
        print(f"    [{base_name}] 第 {i + 1}/{total} 題…", flush=True)
        print(f"      教材子句（前100字）：{clauses_preview or '(無)'}", flush=True)
        logger.info("[%s] 生成第 %d/%d 題", base_name, i + 1, total)
        row = generate_one_question(agent, template)
        if row is None:
            logger.warning("第 %d 題生成失敗，略過", i + 1)
            continue
        results.append(row)
    if not results:
        return 0
    out_name = f"{base_name}_gen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(OUTPUT_HEADERS)
    for idx, r in enumerate(results, 1):
        ws.append([
            idx,
            r.get("stem", ""),
            r.get("option1", ""),
            r.get("option2", ""),
            r.get("option3", ""),
            r.get("option4", ""),
            r.get("answer", ""),
            r.get("章"),
            r.get("節"),
            r.get("頁碼"),
            r.get("難度"),
        ])
    wb.save(out_path)
    print(f"    已寫入 {len(results)} 題至 {out_path}", flush=True)
    return len(results)


def main():
    if not os.path.isdir(INPUT_DIR):
        print(f"輸入目錄不存在：{INPUT_DIR}")
        sys.exit(1)
    xlsx_files = sorted(Path(INPUT_DIR).glob("*.xlsx"))
    if not xlsx_files:
        print(f"目錄內無 xlsx 檔案：{INPUT_DIR}")
        sys.exit(1)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"輸入目錄：{INPUT_DIR}")
    print(f"輸出目錄：{OUTPUT_DIR}")
    print(f"共 {len(xlsx_files)} 個 xlsx 待處理", flush=True)

    print("正在連接 MQTT 與啟動 agent…", flush=True)
    config = app_helper.get_agent_config()
    agent = GenQuestionsAgent(config)
    agent.start_thread()
    time.sleep(1)
    agent.subscribe(agent.agent_id)
    print("開始處理（首次呼叫 LLM 可能需 10～30 秒，屬正常）…", flush=True)

    total_generated = 0
    for fp in xlsx_files:
        file_path = str(fp)
        print(f"\n處理：{fp.name}", flush=True)
        n = process_one_file(agent, file_path, OUTPUT_DIR)
        total_generated += n

    agent.terminate()
    time.sleep(1)
    print(f"\n完成。共產生 {total_generated} 題。", flush=True)


if __name__ == "__main__":
    main()
